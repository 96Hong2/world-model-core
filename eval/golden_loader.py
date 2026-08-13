"""golden.yaml 로더와 원문 대조기.

원문 대조기(`verify_quote`)는 기대값이 자료 원문에서 왔는지를 기계로 확인하는 장치다.
기대값을 시스템 출력에서 베껴 오면 이 대조에서 걸린다.
"""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml

GOLDEN_PATH = Path(__file__).with_name("golden.yaml")

# 원문·답변 비교는 항상 이 정규화를 거친다.
# NFKC 로 전각/반각과 호환 문자를 접고, 제로폭 문자를 지우고, 연속 공백을 하나로 만든다.
_ZERO_WIDTH = dict.fromkeys(map(ord, "​‌‍⁠﻿­"), None)
_WS = re.compile(r"\s+")


def normalize(text: str) -> str:
    if text is None:
        return ""
    t = unicodedata.normalize("NFKC", str(text)).translate(_ZERO_WIDTH)
    return _WS.sub(" ", t).strip()


def contains(haystack: str, needle: str) -> bool:
    return normalize(needle) in normalize(haystack)


@dataclass(frozen=True)
class Golden:
    version: str
    paths: dict[str, str]
    questions: list[dict[str, Any]]

    def by_id(self, qid: str) -> dict[str, Any]:
        for q in self.questions:
            if q["id"] == qid:
                return q
        raise KeyError(f"golden.yaml 에 없는 문항: {qid}")

    def expand(self, value: str) -> str:
        """'{wiki}/...' 같은 별칭 경로를 실제 경로로 바꾼다."""
        out = value
        for key, base in self.paths.items():
            out = out.replace("{" + key + "}", base)
        return out


def load_golden(path: Path | str = GOLDEN_PATH) -> Golden:
    data = yaml.safe_load(Path(path).read_text(encoding="utf-8"))
    return Golden(
        version=data["version"],
        paths=data["paths"],
        questions=data["questions"],
    )


def iter_expected_quotes(question: dict[str, Any]):
    """한 문항이 주장하는 (source_quote, verify) 쌍을 전부 돌려준다."""
    for item in question.get("must_include", []) or []:
        if item.get("source_quote") and item.get("verify"):
            yield item
    conflict = question.get("conflict")
    if conflict:
        for side in conflict.get("sides", []):
            if side.get("source_quote") and side.get("verify"):
                yield side


# ---------------------------------------------------------------------------
# 원문 열람기 — 형식별로 딱 필요한 만큼만 읽는다
# ---------------------------------------------------------------------------

_CELL_RE = re.compile(r"^([A-Z]+)(\d+)$")


def _col_to_index(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _read_xlsx_cell(path: Path, sheet: str, cell: str) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet]
        m = _CELL_RE.match(cell)
        if not m:
            raise ValueError(f"셀 표기가 아니다: {cell}")
        col, row = _col_to_index(m.group(1)), int(m.group(2))
        return "" if (v := ws.cell(row, col).value) is None else str(v)
    finally:
        wb.close()


def _read_xlsx_sheet(path: Path, sheet: str) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet]
        parts = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        return "\n".join(parts)
    finally:
        wb.close()


def _read_slack_message(path: Path, ts: str) -> str:
    with path.open(encoding="utf-8") as fh:
        for line in fh:
            if not line.strip():
                continue
            rec = json.loads(line)
            if rec.get("ts") == ts:
                return rec.get("text") or ""
    raise LookupError(f"{path.name} 에 ts={ts} 메시지가 없다")


def read_source_text(golden: Golden, verify: dict[str, Any], source_file: str) -> str:
    """verify 블록이 가리키는 원문 조각을 그대로 읽어 온다."""
    kind = verify["kind"]

    if kind == "xlsx":
        path = Path(golden.expand(source_file))
        if verify.get("mode") == "sheet_contains":
            return _read_xlsx_sheet(path, verify["sheet"])
        return _read_xlsx_cell(path, verify["sheet"], verify["cell"])

    if kind == "text":
        # PDF·PPTX·HTML 원본은 그대로 읽을 수 없으므로 A2 이전 단계에서 뽑아 둔
        # 추출 txt 로 대조한다. 어떤 파일에서 뽑았는지는 source_file 이 남긴다.
        return Path(golden.expand(verify["path"])).read_text(encoding="utf-8")

    if kind == "slack":
        return _read_slack_message(Path(golden.expand(verify["path"])), verify["ts"])

    raise ValueError(f"모르는 verify.kind: {kind}")


def verify_quote(golden: Golden, item: dict[str, Any]) -> tuple[bool, str]:
    """source_quote 가 원문의 그 자리에 실제로 있는지 확인한다."""
    try:
        raw = read_source_text(golden, item["verify"], item["source_file"])
    except (FileNotFoundError, LookupError, KeyError) as exc:
        return False, f"원문을 열지 못함: {exc}"
    if contains(raw, item["source_quote"]):
        return True, ""
    return False, (
        f"원문에 없는 인용: {item['source_quote']!r} @ {item.get('locator')} "
        f"(원문 앞부분: {normalize(raw)[:120]!r})"
    )
