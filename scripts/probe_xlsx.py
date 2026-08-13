#!/usr/bin/env python
"""R0 자료 실측용 임시 probe 스크립트.

data/sources/wiki 안의 xlsx 8종 구조를 읽기 전용으로 조사한다.
원본은 절대 수정하지 않는다(openpyxl load_workbook 후 save 하지 않음).

usage:
  python scripts/probe_xlsx.py overview            # 전 파일 시트/치수/병합셀 개요
  python scripts/probe_xlsx.py head <파일키> [n]    # 시트별 상단 n행 덤프
  python scripts/probe_xlsx.py cols <파일키> <시트> <헤더행>  # 컬럼별 값 예시+enum 빈도
"""
import sys
import datetime
from collections import Counter, OrderedDict
from pathlib import Path

import openpyxl

SRC = Path("data/sources/wiki")

FILES = OrderedDict([
    ("featuremap", "오로라웍스 기능맵 v2.1.xlsx"),
    ("activity", "영업본부_방문 및 접촉 활동 일지.xlsx"),
    ("weekly", "영업본부_주간활동계획.xlsx"),
    ("bd", "BD Overview.xlsx"),
    ("painpoint", "고객사 Pain Point·구축 대응 통합 관리.xlsx"),
    ("compliance", "채권추심-컴플라이언스 위반 체크리스트.xlsx"),
    ("agenttest", "채권추심 지원 Agent 테스트.xlsx"),
    ("aiqa", "AI QA 대화 예제.xlsx"),
])


def load(key, data_only=True):
    return openpyxl.load_workbook(SRC / FILES[key], data_only=data_only, read_only=False)


def cell_repr(v, limit=120):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"<{type(v).__name__}:{v.isoformat()}>"
    s = str(v).replace("\n", "\\n").replace("\r", "")
    return s[:limit] + ("…" if len(s) > limit else "")


def used_bounds(ws):
    """실제 값이 있는 마지막 행/열을 센다(openpyxl dims는 서식만 있어도 부풀려진다)."""
    max_r = 0
    max_c = 0
    nonempty_rows = 0
    for r_idx, row in enumerate(ws.iter_rows(), start=1):
        row_has = False
        for c_idx, c in enumerate(row, start=1):
            if c.value is not None and str(c.value).strip() != "":
                row_has = True
                if c_idx > max_c:
                    max_c = c_idx
        if row_has:
            max_r = r_idx
            nonempty_rows += 1
    return max_r, max_c, nonempty_rows


def cmd_overview():
    for key, fname in FILES.items():
        p = SRC / fname
        print("=" * 100)
        print(f"[{key}] {fname}  ({p.stat().st_size:,} bytes)")
        wb = load(key)
        print(f"  sheets({len(wb.sheetnames)}): {wb.sheetnames}")
        for ws in wb.worksheets:
            mr, mc, ner = used_bounds(ws)
            merged = list(ws.merged_cells.ranges)
            print(f"  - '{ws.title}' state={ws.sheet_state} dims={ws.dimensions} "
                  f"declared={ws.max_row}x{ws.max_column} used={mr}x{mc} nonempty_rows={ner} "
                  f"merged={len(merged)}")
            if merged:
                print(f"      merged sample: {[str(m) for m in merged[:12]]}")
        wb.close()


def cmd_head(key, n=12, sheet=None):
    wb = load(key)
    for ws in wb.worksheets:
        if sheet and ws.title != sheet:
            continue
        mr, mc, _ = used_bounds(ws)
        print("=" * 100)
        print(f"[{key}] sheet='{ws.title}' used={mr}x{mc}")
        for r in range(1, min(int(n), mr) + 1):
            vals = []
            for c in range(1, min(mc, 30) + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() != "":
                    vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}={cell_repr(v, 90)}")
            if vals:
                print(f"  r{r}: " + " | ".join(vals))
    wb.close()


def cmd_cols(key, sheet, header_row, samples=3, enum_max=60):
    samples, enum_max = int(samples), int(enum_max)
    wb = load(key)
    ws = wb[sheet]
    hr = int(header_row)
    mr, mc, _ = used_bounds(ws)
    print("=" * 100)
    print(f"[{key}] '{sheet}' header_row={hr} used={mr}x{mc}")
    for c in range(1, mc + 1):
        header = cell_repr(ws.cell(row=hr, column=c).value, 60)
        col_vals = []
        types = Counter()
        for r in range(hr + 1, mr + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or str(v).strip() == "":
                continue
            col_vals.append(v)
            types[type(v).__name__] += 1
        cnt = Counter(str(v).strip() for v in col_vals)
        uniq = len(cnt)
        letter = openpyxl.utils.get_column_letter(c)
        print(f"\n  [{letter}] '{header}' filled={len(col_vals)} unique={uniq} types={dict(types)}")
        ex = [cell_repr(v, 100) for v in col_vals[:samples]]
        print(f"      ex: {ex}")
        if uniq <= enum_max:
            print(f"      ENUM({uniq}): " + " / ".join(f"{k}({n})" for k, n in cnt.most_common()))
    wb.close()


def cmd_colvals(key, sheet, header_row, col_letter, limit=500):
    """한 컬럼의 값 전체를 순서대로 덤프(중복 포함/제거 둘 다)."""
    wb = load(key)
    ws = wb[sheet]
    hr = int(header_row)
    mr, _, _ = used_bounds(ws)
    ci = openpyxl.utils.column_index_from_string(col_letter)
    vals = []
    for r in range(hr + 1, mr + 1):
        v = ws.cell(row=r, column=ci).value
        vals.append(None if v is None else str(v).strip())
    cnt = Counter(v for v in vals if v)
    print(f"[{key}] '{sheet}' col {col_letter} rows={len(vals)} blank={sum(1 for v in vals if not v)} unique={len(cnt)}")
    for k, n in cnt.most_common(limit):
        print(f"  {n:5d}  {k}")
    wb.close()


def cmd_dump(key, sheet, r1, r2, c2=30):
    wb = load(key)
    ws = wb[sheet]
    for r in range(int(r1), int(r2) + 1):
        vals = []
        for c in range(1, int(c2) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                vals.append(f"{openpyxl.utils.get_column_letter(c)}={cell_repr(v, 300)}")
        print(f"r{r}: " + " | ".join(vals))
    wb.close()


if __name__ == "__main__":
    cmd = sys.argv[1]
    args = sys.argv[2:]
    {"overview": cmd_overview, "head": cmd_head, "cols": cmd_cols,
     "colvals": cmd_colvals, "dump": cmd_dump}[cmd](*args)
