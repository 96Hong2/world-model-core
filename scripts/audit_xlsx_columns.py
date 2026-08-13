#!/usr/bin/env python
"""xlsx 발췌 커버리지 감사 — 한 행의 값이 그 행 발췌에 전부 실리는가.

배경: 행을 앞쪽 컬럼 한 덩어리로 묶어 발췌를 만들면 뒤쪽 컬럼 값이 excerpt 에서 사라진다.
사라진 값은 Evidence fulltext 검색에서 영영 안 잡힌다(structured 에만 남는다).

측정 방법은 전역 대조가 아니라 **행 단위 대조**다. Evidence 레코드마다 `structured.sheet`·`row`
가 있으므로 원본의 같은 행을 다시 읽어 셀 하나씩 그 Evidence 의 excerpt 와 맞춰 본다.
전역 대조는 헤더 라벨·날짜 표기·다른 행의 같은 값 때문에 판정이 흐려진다.

판정 5종:
  covered  발췌에 실렸다
  masked   PII 라 마스킹됐다(발췌에 없는 것이 정상). 열 이름 마스킹(NE-COLUMN)도 여기다
  trivial  ✔·- 같은 표기 기호
  merged   Evidence 가 안 붙은 행인데 값이 앞뒤 행 발췌에 병합돼 들어갔다
  missing  값이 있는데 어디에도 없다 ← 이것만 결함

행 단위 대조와 별도로, 값이 있고 Evidence 도 없고 이웃 행에도 병합되지 않은 행을 센다(uncollected).

읽기 전용이다. 원본 xlsx 를 저장하지 않고 data/parsed 도 고치지 않는다.

usage:
  python scripts/audit_xlsx_columns.py                  # 전 xlsx 소스
  python scripts/audit_xlsx_columns.py src_bd_overview  # 소스 하나
  python scripts/audit_xlsx_columns.py --json
  python scripts/audit_xlsx_columns.py --max-missing 0  # 게이트로 쓸 때
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import OrderedDict, defaultdict
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))

from openpyxl.utils import get_column_letter  # noqa: E402

from ingestion.masking import default_masker  # noqa: E402
from ingestion.parsers.xlsx_common import (  # noqa: E402
    PARSED_DIR,
    Sheet,
    flat_text,
    load_workbook,
    parse_cell_date,
    source_config,
)

XLSX_SOURCES = [
    "src_featuremap_v21",
    "src_sales_activity_log",
    "src_sales_weekly_plan",
    "src_bd_overview",
    "src_pain_registry",
]

# 값이 있어도 발췌에 없어도 되는 표기 기호. 뜻은 범례 Evidence 가 따로 담는다.
TRIVIAL_VALUES = {
    "✔", "○", "●", "◎", "△", "▲", "■", "□", "-", "–", "—",
    "N/A", "n/a", "O", "X", "o", "x", "v", "V", "<->", "↔", "→",
}
# 발췌에 없는 것이 맞다고 판단한 셀. 이유를 함께 적는다.
# 여기 없는 새 누락이 생기면 감사가 빨간불이 되게 하려고 예외를 목록으로 못 박는다.
KNOWN_MISSING: dict[str, str] = {
    # 주간활동계획 이영희 시트는 같은 행에 표 두 개가 가로로 나란히 놓여 있다.
    # 왼쪽(C~H)은 '5.·6. 고객사 표', 오른쪽(L~R)은 '26년 월별 매출계획/계약계획' 이다.
    # 오른쪽 값을 왼쪽 행 발췌에 붙이면 그 고객사가 그 금액을 계획한 것처럼 읽혀
    # 없는 사실을 만든다. 붙이지 않는 것이 맞다.
    # 오른쪽 표 자체는 config/sources.yaml 의 tables_include 밖이다(Lead 판단 필요).
    "이영희!L24": "월별계획 표(다른 표) 값 — 고객사 행 발췌에 붙이면 오독",
    "이영희!M24": "월별계획 표(다른 표) 값 — 고객사 행 발췌에 붙이면 오독",
    "이영희!N24": "월별계획 표(다른 표) 값 — 고객사 행 발췌에 붙이면 오독",
    "이영희!N30": "월별계획 표(다른 표) 값 — 고객사 행 발췌에 붙이면 오독",
    "이영희!N31": "월별계획 표(다른 표) 값 — 고객사 행 발췌에 붙이면 오독",
    "이영희!N32": "월별계획 표(다른 표) 값 — 고객사 행 발췌에 붙이면 오독",
}

_NUM_RE = re.compile(r"^-?[\d,]+(?:\.\d+)?$")
# 발췌는 EXCERPT_MAX(500)에서 잘린다. 500자를 넘는 셀은 앞부분만 비교한다.
PREFIX_COMPARE = 40


def _variants(text: str, raw) -> list[str]:
    """같은 값의 다른 표기들. 파서가 날짜·금액을 사람이 읽는 형태로 바꿔 적기 때문이다."""
    out = [text]
    iso, status = parse_cell_date(raw)
    if status == "ok" and iso:
        out.append(iso)
    if _NUM_RE.match(text):
        bare = text.replace(",", "")
        out.append(bare)
        try:
            out.append(f"{int(float(bare)):,}")
        except ValueError:
            pass
    if len(text) > PREFIX_COMPARE:
        out.append(text[:PREFIX_COMPARE])
    return [v for v in out if v]


def _evidence_rows(
    source_id: str, parsed_dir: Path | None = None
) -> tuple[dict[tuple[str, int], list[dict]], set[str]]:
    """(시트, 행) → 그 행에서 나온 Evidence 목록. 두 번째 값은 Evidence 가 있는 시트 집합."""
    path = (parsed_dir or PARSED_DIR) / f"{source_id}.jsonl"
    if not path.exists():
        raise SystemExit(f"{path} 가 없다. 먼저 파서를 돌린다")
    by_row: dict[tuple[str, int], list[dict]] = defaultdict(list)
    sheets: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        st = rec.get("structured") or {}
        sheet, row = st.get("sheet"), st.get("row")
        if not sheet or not isinstance(row, int):
            continue
        # 시트 이름에 '(각주)'·'(범례)' 같은 꼬리가 붙는 경우가 있다.
        base = sheet.split("(각주")[0].split("(범례")[0]
        by_row[(base, row)].append(rec)
        sheets.add(base)
    return by_row, sheets


def _haystack(recs) -> str:
    return "\n".join(
        (r.get("excerpt") or "")
        + "\n"
        + "\n".join((r.get("structured") or {}).get("raw_cells", {}).values())
        for r in (recs or [])
    )


def _ne_column_map(sheet: Sheet, last_col: int, masker) -> set[int]:
    """열 이름 마스킹(NE-COLUMN) 대상 컬럼. 헤더 라벨이 config 목록에 있는 열이다.

    이 열의 값은 마스킹돼 발췌에 원문이 없는 것이 정상이다. 값만 보고는 인명인지
    알 수 없으므로(정규식으로 한국어 인명을 못 잡는다) 파서와 같은 기준인 열 이름으로 본다.

    시트 상단만 보면 안 된다. 일일보고 시트는 섹션 헤더가 40행 아래에도 있어서
    '6. 전략 BD 활동' 의 `담당자` 열을 놓친다(실측 4건이 누락으로 잘못 잡혔다).
    """
    labels = {str(x).strip() for x in getattr(masker._rules, "column_labels", ())}
    if not labels:
        return set()
    out: set[int] = set()
    for row in range(1, sheet.max_row + 1):
        for c in range(1, last_col + 1):
            if (sheet.flat(row, c) or "") in labels:
                out.add(c)
    return out


def audit_source(source_id: str, parsed_dir: Path | None = None) -> dict:
    cfg = source_config(source_id)
    by_row, parsed_sheets = _evidence_rows(source_id, parsed_dir)
    masker = default_masker()
    wb = load_workbook(Path(cfg["path"]))

    cols: dict[tuple[str, int], dict] = {}
    uncollected: dict[str, list[int]] = defaultdict(list)
    skipped_sheets: list[str] = []

    for ws_raw in wb.worksheets:
        sheet = Sheet(ws_raw)
        if sheet.title not in parsed_sheets:
            skipped_sheets.append(sheet.title)
            continue
        last_col = min(ws_raw.max_column or 1, 40)
        masked_cols = _ne_column_map(sheet, last_col, masker)
        # 시트 전체 발췌. Evidence 가 안 붙은 행의 값이 이웃 행으로 병합됐는지 볼 때 쓴다.
        sheet_hay = "\n".join(
            _haystack(recs) for (s, _r), recs in by_row.items() if s == sheet.title
        )
        for row in range(1, sheet.max_row + 1):
            recs = by_row.get((sheet.title, row))
            values = [(c, sheet.flat(row, c)) for c in range(1, last_col + 1)]
            values = [(c, v) for c, v in values if v]
            if not values:
                continue
            hay = _haystack(recs) if recs else sheet_hay
            for c, text in values:
                slot = cols.setdefault(
                    (sheet.title, c),
                    {
                        "covered": 0, "masked": 0, "trivial": 0,
                        "merged": 0, "orphan": 0, "known": 0, "missing": 0, "samples": [],
                    },
                )
                if text in TRIVIAL_VALUES or len(text) < 2:
                    slot["trivial"] += 1
                    continue
                raw = sheet.raw(row, c)
                if any(v in hay for v in _variants(text, raw)):
                    slot["covered" if recs else "merged"] += 1
                    continue
                # 마스킹 대상이면 발췌에 없는 것이 정상이다.
                # 열 이름 마스킹(NE-COLUMN)은 값 자체로는 판별할 수 없어 열로 본다.
                if c in masked_cols or masker.mask(text).hits:
                    slot["masked"] += 1
                    continue
                if not recs:
                    # Evidence 가 안 붙은 행이다. 컬럼 누락이 아니라 '행 미수집'이라
                    # missing 으로 세지 않는다. 아래 uncollected 에 행 번호로 남는다.
                    slot["orphan"] += 1
                    continue
                if sheet.locator(row, c) in KNOWN_MISSING:
                    slot["known"] += 1
                    continue
                slot["missing"] += 1
                if len(slot["samples"]) < 5:
                    slot["samples"].append(
                        {"cell": sheet.locator(row, c), "value": text[:120],
                         "has_evidence": bool(recs)}
                    )
            if not recs:
                uncollected[sheet.title].append(row)
    wb.close()

    sheets_report: dict[str, dict] = OrderedDict()
    for (sheet, c) in sorted(cols):
        slot = cols[(sheet, c)]
        sheets_report.setdefault(sheet, {"columns": OrderedDict(), "uncollected_rows": []})
        sheets_report[sheet]["columns"][get_column_letter(c)] = slot
    for sheet, rows in uncollected.items():
        sheets_report.setdefault(sheet, {"columns": OrderedDict(), "uncollected_rows": []})
        sheets_report[sheet]["uncollected_rows"] = rows
    return {
        "source_id": source_id,
        "path": cfg["path"],
        "sheets": sheets_report,
        "sheets_without_evidence": skipped_sheets,
    }


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("sources", nargs="*", default=None)
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--max-missing", type=int, default=None,
                    help="누락 셀이 이 값을 넘으면 exit 1")
    ap.add_argument("--show-uncollected", action="store_true")
    ap.add_argument("--parsed-dir", default=None,
                    help="data/parsed 대신 볼 디렉토리(감사기 자체 검증용)")
    args = ap.parse_args(argv)

    parsed_dir = Path(args.parsed_dir) if args.parsed_dir else None
    reports = [audit_source(n, parsed_dir) for n in (args.sources or XLSX_SOURCES)]
    if args.json:
        print(json.dumps(reports, ensure_ascii=False, indent=2))
    total = 0
    for rep in reports:
        if not args.json:
            print(f"\n=== {rep['source_id']} ===")
        for sheet, data in rep["sheets"].items():
            bad = {k: v for k, v in data["columns"].items() if v["missing"] > 0}
            total += sum(v["missing"] for v in bad.values())
            if args.json:
                continue
            checked = sum(
                v["covered"] + v["masked"] + v["merged"] + v["known"] + v["missing"]
                for v in data["columns"].values()
            )
            unc = data["uncollected_rows"]
            tail = f" · Evidence 없는 행 {len(unc)}" if unc else ""
            if not bad:
                print(f"  [{sheet}] 전 컬럼 발췌 반영 (검사 {checked}칸){tail}")
            else:
                print(f"  [{sheet}] 검사 {checked}칸{tail}")
                for col, v in bad.items():
                    print(
                        f"    {col}열: 반영 {v['covered']} · 병합 {v['merged']} · "
                        f"마스킹 {v['masked']} · 알려진예외 {v['known']} · 누락 {v['missing']}"
                    )
                    for s in v["samples"][:3]:
                        print(f"        {s['cell']} = {s['value']}")
            if unc and args.show_uncollected:
                print(f"      Evidence 없는 행: {unc[:20]}{' …' if len(unc) > 20 else ''}")
        if not args.json and rep["sheets_without_evidence"]:
            print(f"  (Evidence 0건 시트: {', '.join(rep['sheets_without_evidence'])})")
    if not args.json:
        print(f"\n총 누락 셀 {total}칸")
    if args.max_missing is not None and total > args.max_missing:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
