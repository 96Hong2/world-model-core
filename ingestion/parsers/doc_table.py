"""표 형태 xlsx 를 좌표 그대로 읽는 범용 파서.

기능맵·활동일지·BD Overview·Pain 레지스트리처럼 스키마가 확정된 xlsx 는 전용 파서가 따로 있다.
이 파서는 그 밖의 자료(AI QA 대화 예제, 컴플라이언스 체크리스트, Agent 테스트 결과)를 위한 것이다.
스키마를 짐작하지 않고 헤더 행만 찾아 "컬럼명: 값" 으로 펼친다.

들어 있는 규칙 셋:
  - 헤더는 앞쪽 행 중 값이 2칸 이상인 첫 행이다. 못 찾으면 컬럼 문자(A·B·C)를 이름으로 쓴다.
  - 같은 내용의 행은 한 번만 남긴다. 한 파일 안에 시트가 통째로 복사돼 있는 경우가 실재한다
    (`AI QA 대화 예제.xlsx` 의 '채권추심-컴플라이언스위반' ↔ '로우데이터').
  - 세로 병합은 아래로 채운다. 가로 병합은 건드리지 않는다.
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from ingestion.parsers.doc_common import (
    Unit,
    build_evidences,
    build_source_record,
    normalize_text,
)

HEADER_SEARCH_ROWS = 10
MIN_HEADER_CELLS = 2
MAX_COLUMNS = 30
LONG_LABEL = 40  # 이보다 긴 첫 행은 헤더가 아니라 안내 문장이다


@dataclass(frozen=True)
class TableRow:
    sheet: str
    row: int
    text: str


def _cell_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = re.sub(r"[ \t]+", " ", str(value)).strip()
    return text or None


def _vertical_fill(ws, last_col: int) -> dict[tuple[int, int], object]:
    filled: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        if rng.max_row <= rng.min_row or rng.min_col > last_col:
            continue
        top = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, min(rng.max_col, last_col) + 1):
                filled[(r, c)] = top
    return filled


def find_header_row(ws, last_col: int) -> int | None:
    for row in range(1, min(ws.max_row, HEADER_SEARCH_ROWS) + 1):
        labels = [_cell_text(ws.cell(row, c).value) for c in range(1, last_col + 1)]
        present = [lbl for lbl in labels if lbl]
        if len(present) >= MIN_HEADER_CELLS and all(len(lbl) <= LONG_LABEL for lbl in present):
            return row
    return None


def read_table(ws) -> list[TableRow]:
    last_col = min(ws.max_column or 1, MAX_COLUMNS)
    header_row = find_header_row(ws, last_col)
    fill = _vertical_fill(ws, last_col)

    def value(row: int, col: int) -> str | None:
        raw = ws.cell(row, col).value
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            raw = fill.get((row, col))
        return _cell_text(raw)

    if header_row is None:
        labels = {c: get_column_letter(c) for c in range(1, last_col + 1)}
        start = 1
    else:
        labels = {}
        for c in range(1, last_col + 1):
            labels[c] = value(header_row, c) or get_column_letter(c)
        start = header_row + 1

    rows: list[TableRow] = []
    for row in range(start, ws.max_row + 1):
        parts = []
        for c in range(1, last_col + 1):
            v = value(row, c)
            if v:
                parts.append(f"{labels[c]}: {v}")
        if not parts:
            continue
        rows.append(TableRow(sheet=ws.title, row=row, text=" | ".join(parts)))
    return rows


def _units(rows: list[TableRow]) -> list[Unit]:
    """내용이 같은 행은 처음 것만 남긴다. 시트 통째 복사본을 두 벌 넣지 않기 위해서다."""
    seen: set[str] = set()
    units: list[Unit] = []
    for item in rows:
        text = normalize_text(item.text)
        if not text:
            continue
        digest = hashlib.sha256(re.sub(r"\s+", "", text).encode("utf-8")).hexdigest()
        if digest in seen:
            continue
        seen.add(digest)
        units.append(
            Unit(
                locator=f"{item.sheet}!R{item.row}",
                unit="row",
                text=text,
                structured={
                    "record_kind": "doc_section",
                    "sheet": item.sheet,
                    "row": item.row,
                    "섹션": item.sheet,
                },
            )
        )
    return units


def parse_source(cfg: dict) -> tuple[dict, list[dict]]:
    scope = cfg.get("scope") or {}
    include = set(scope.get("sheets_include") or [])
    exclude = set(scope.get("sheets_exclude") or [])

    wb = openpyxl.load_workbook(Path(cfg["path"]), data_only=True)
    rows: list[TableRow] = []
    used_sheets: list[str] = []
    for ws in wb.worksheets:
        if include and ws.title not in include:
            continue
        if ws.title in exclude:
            continue
        used_sheets.append(ws.title)
        rows.extend(read_table(ws))
    wb.close()

    units = _units(rows)
    evidences = build_evidences(cfg["id"], units)
    source = build_source_record(
        cfg,
        parser="doc_table.xlsx",
        extractor="openpyxl",
        evidence_count=len(evidences),
        notes=(
            f"시트 {len(used_sheets)}개({', '.join(used_sheets)}) · 원본 행 {len(rows)} → "
            f"중복 제거 후 {len(units)}"
        ),
    )
    return source, evidences
