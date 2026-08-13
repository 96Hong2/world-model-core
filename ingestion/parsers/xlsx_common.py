"""xlsx 파서 공통 부품.

전부 deterministic 이다. LLM 호출이 없다.
원본 xlsx 는 열기만 하고 저장하지 않는다.
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import json
import re
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl
import yaml
from jsonschema import Draft202012Validator
from openpyxl.utils import get_column_letter

from ingestion.masking import MaskedText, Masker, default_masker, merge_hits

REPO = Path(__file__).resolve().parents[2]
CONFIG_DIR = REPO / "config"
CONTRACTS_DIR = REPO / "contracts"
PARSED_DIR = REPO / "data" / "parsed"

PARSER_VERSION = "1.0.0"
EXCERPT_MAX = 500

# 파이썬이 줄바꿈으로 세지만 JSON 은 문자열 안 평범한 문자로 취급하는 것들.
# 그대로 두면 jsonl 한 줄이 str.splitlines() 에서 두 줄로 쪼개져 적재가 통째로 깨진다.
EXOTIC_LINE_BREAKS = str.maketrans({c: "\n" for c in "\x0b\x0c\x1c\x1d\x1e\x85  "})

_WHITESPACE_RUN = re.compile(r"\s+")

# contracts/parsed_record.schema.json 의 sensitivity 는 low/medium/high 이고
# config/sources.yaml 은 internal/restricted 로 적는다. 계약 쪽 값으로 옮긴다.
SENSITIVITY_MAP = {
    "public": "low",
    "internal": "medium",
    "restricted": "high",
    "low": "low",
    "medium": "medium",
    "high": "high",
}
# 같은 이유로 doc_status 도 계약 enum(final/draft/wip/deprecated)으로 옮긴다.
DOC_STATUS_MAP = {
    "approved": "final",
    "final": "final",
    "draft": "draft",
    "wip": "wip",
    "deprecated": "deprecated",
}


# --------------------------------------------------------------------------
# 결과 컨테이너
# --------------------------------------------------------------------------
@dataclass
class ParseResult:
    source: dict
    evidence: list[dict] = field(default_factory=list)
    stats: dict = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def add_failure(self, n: int = 1) -> None:
        self.stats["date_parse_failures"] = self.stats.get("date_parse_failures", 0) + n


# --------------------------------------------------------------------------
# config 읽기
# --------------------------------------------------------------------------
@lru_cache(maxsize=None)
def _sources_config() -> dict:
    return yaml.safe_load((CONFIG_DIR / "sources.yaml").read_text(encoding="utf-8"))


def source_config(source_id: str) -> dict:
    from ingestion.source_registry import expand_config

    for s in expand_config(_sources_config()):
        if s["id"] == source_id:
            return s
    raise KeyError(f"config/sources.yaml 에 {source_id} 가 없다")


@lru_cache(maxsize=None)
def bd_seed() -> dict:
    return yaml.safe_load((CONFIG_DIR / "bd-seed.yaml").read_text(encoding="utf-8"))


@lru_cache(maxsize=None)
def _schema() -> dict:
    return json.loads((CONTRACTS_DIR / "parsed_record.schema.json").read_text(encoding="utf-8"))


def validate_records(source: dict, evidence: Sequence[dict]) -> list[str]:
    """계약 스키마 위반을 문자열 목록으로 돌려준다(빈 목록 = 통과)."""
    schema = _schema()
    src_validator = Draft202012Validator({**schema, "$ref": "#/$defs/SourceRecord"})
    ev_validator = Draft202012Validator({**schema, "$ref": "#/$defs/EvidenceRecord"})
    errors: list[str] = []
    for e in src_validator.iter_errors(source):
        errors.append(f"source: {list(e.path)} {e.message}")
    for i, rec in enumerate(evidence):
        for e in ev_validator.iter_errors(rec):
            errors.append(f"evidence[{i}] {rec.get('locator')}: {list(e.path)} {e.message}")
    return errors


# --------------------------------------------------------------------------
# 해시 / id
# --------------------------------------------------------------------------
def flat_text(value: Any) -> str | None:
    """셀 안 줄바꿈과 연속 공백을 공백 하나로 접는다.

    Excel 셀은 Alt+Enter 줄바꿈을 흔하게 쓴다(진척도!D2 = "세일즈킷\\n완성").
    접지 않으면 "세일즈킷 완성" 으로 찾는 발췌 검색이 영영 안 잡힌다.
    접기 전 값은 EvidenceBuilder 가 structured.raw_cells 에 남겨 원문을 복원할 수 있다.
    """
    if value is None:
        return None
    return _WHITESPACE_RUN.sub(" ", str(value)).strip() or None


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


_ALPHA = "abcdefghijklmnopqrstuvwxyz"


def alpha_digest(text: str, length: int = 24) -> str:
    """숫자가 하나도 없는 sha256 다이제스트.

    16진수 다이제스트는 Evidence 레코드와 같이 실려 나가는데, 우연히 전화번호·주민번호
    정규식에 걸린다(실측: 기능맵 459건에서 58회). 식별자 때문에 마스킹 게이트가 빨간불이
    되는 것을 막으려고 자모만 쓴다. 원본은 여전히 sha256 이라 결정성은 그대로다.
    """
    n = int(hashlib.sha256(text.encode("utf-8")).hexdigest(), 16)
    out = []
    for _ in range(length):
        n, r = divmod(n, 26)
        out.append(_ALPHA[r])
    return "".join(out)


def make_evidence_id(source_id: str, locator: str) -> str:
    """source_id + locator 로 결정적으로 만든다. 2회 파싱 시 같은 값이 나온다."""
    return f"ev_{alpha_digest(f'{source_id}|{locator}')}"


# --------------------------------------------------------------------------
# 날짜
# --------------------------------------------------------------------------
_EXCEL_EPOCH = _dt.date(1899, 12, 30)  # Excel 의 1900 윤년 버그를 포함한 기준일
_DATE_TEXT_RES = (
    re.compile(r"^(\d{4})\D{1,3}(\d{1,2})\D{1,3}(\d{1,2})"),
    re.compile(r"^(\d{4})(\d{2})(\d{2})$"),
)


def excel_serial_to_date(serial: float) -> _dt.date:
    return _EXCEL_EPOCH + _dt.timedelta(days=int(serial))


def parse_cell_date(value: Any) -> tuple[str | None, str]:
    """셀 값을 ISO 날짜로 바꾼다.

    돌려주는 상태는 ok / empty / unparsable 셋이다.
    '-' 처럼 값 없음을 뜻하는 표기는 실패가 아니라 empty 다.
    """
    if value is None:
        return None, "empty"
    if isinstance(value, _dt.datetime):
        return value.date().isoformat(), "ok"
    if isinstance(value, _dt.date):
        return value.isoformat(), "ok"
    if isinstance(value, (int, float)):
        # Excel serial. 실측상 대상 파일에는 없지만 들어오면 변환한다.
        if 20000 <= float(value) <= 80000:
            return excel_serial_to_date(value).isoformat(), "ok"
        return None, "unparsable"
    text = str(value).strip()
    if text == "" or text in {"-", "–", "—", "N/A", "n/a"}:
        return None, "empty"
    for rx in _DATE_TEXT_RES:
        m = rx.match(text)
        if m:
            y, mo, d = (int(g) for g in m.groups())
            try:
                return _dt.date(y, mo, d).isoformat(), "ok"
            except ValueError:
                return None, "unparsable"
    return None, "unparsable"


# --------------------------------------------------------------------------
# 워크북 / 병합셀
# --------------------------------------------------------------------------
def load_workbook(path: Path):
    """읽기 전용으로 연다. 절대 save 하지 않는다."""
    return openpyxl.load_workbook(path, data_only=True)


def vertical_fill_map(ws, columns: Iterable[int]) -> dict[tuple[int, int], Any]:
    """세로 병합만 아래로 채운다.

    가로 병합은 건드리지 않는다. 방향을 구분하지 않고 채우면 서술 텍스트가 옆 컬럼으로 번진다.
    """
    cols = set(columns)
    filled: dict[tuple[int, int], Any] = {}
    for rng in ws.merged_cells.ranges:
        if rng.max_row <= rng.min_row:
            continue  # 세로 방향이 아니면 대상이 아니다
        if rng.min_col not in cols:
            continue
        top = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if c in cols:
                    filled[(r, c)] = top
    return filled


class Sheet:
    """한 시트를 좌표로 읽는 얇은 래퍼. 세로 병합 ffill 을 알고 있다."""

    def __init__(self, ws, ffill_columns: Sequence[int] = ()):
        self.ws = ws
        self.title = ws.title
        self.state = ws.sheet_state
        self._fill = vertical_fill_map(ws, ffill_columns) if ffill_columns else {}

    @property
    def max_row(self) -> int:
        return self.ws.max_row

    def raw(self, row: int, col: int) -> Any:
        v = self.ws.cell(row, col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            return self._fill.get((row, col))
        return v

    def text(self, row: int, col: int) -> str | None:
        v = self.raw(row, col)
        if v is None:
            return None
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        s = str(v).translate(EXOTIC_LINE_BREAKS).strip()
        return s or None

    def flat(self, row: int, col: int) -> str | None:
        """줄바꿈을 접은 셀 값. 마스킹이 필요 없는 라벨·헤더를 읽을 때 쓴다."""
        return flat_text(self.text(row, col))

    def is_numeric(self, row: int, col: int) -> bool:
        return isinstance(self.raw(row, col), (int, float))

    def locator(self, row: int, col: int) -> str:
        return f"{self.title}!{get_column_letter(col)}{row}"

    def row_has_value(self, row: int, cols: Iterable[int]) -> bool:
        return any(self.text(row, c) for c in cols)

    def row_has_own_value(self, row: int, cols: Iterable[int]) -> bool:
        """병합 ffill 로 채워진 값은 빼고, 그 행이 스스로 가진 값이 있는지 본다.

        세로 병합이 데이터 끝보다 아래까지 뻗어 있으면 ffill 만 보고 빈 행을 데이터로 착각한다.
        """
        for c in cols:
            v = self.ws.cell(row, c).value
            if v is not None and str(v).strip():
                return True
        return False


def header_map(sheet: Sheet, row: int, last_col: int = 26) -> dict[str, int]:
    """헤더 행을 읽어 라벨 → 컬럼 번호 사전을 만든다.

    컬럼 위치를 상수로 두면 안 되는 시트가 있어(같은 G열이 주차마다 다른 뜻) 매번 다시 읽는다.
    """
    out: dict[str, int] = {}
    for c in range(1, last_col + 1):
        label = sheet.text(row, c)
        if label:
            key = re.sub(r"\s+", " ", label).strip()
            out.setdefault(key, c)
    return out


def find_col(hdr: dict[str, int], *names: str) -> int | None:
    for n in names:
        if n in hdr:
            return hdr[n]
    for n in names:
        for k, v in hdr.items():
            if k.startswith(n):
                return v
    return None


# --------------------------------------------------------------------------
# Evidence 조립
# --------------------------------------------------------------------------
class EvidenceBuilder:
    """마스킹된 값만 받아 Evidence 레코드를 만든다."""

    def __init__(self, source_id: str, masker: Masker | None = None):
        self.source_id = source_id
        self.masker = masker or default_masker()
        self._chunks: list[MaskedText] = []
        self._raw_cells: dict[str, str] = {}

    # -- 셀 읽기 (읽는 즉시 마스킹한다) --------------------------------
    def cell(self, sheet: Sheet, row: int, col: int, *, column_name: str | None = None) -> str | None:
        """마스킹 후 줄바꿈을 접은 값을 돌려준다.

        접기 전 값이 다르면 `raw_cells[locator]` 에 남겨 원문을 복원할 수 있게 한다.
        """
        value = sheet.text(row, col)
        if value is None:
            return None
        masked = self.masker.mask(
            value, column_name=column_name, numeric_cell=sheet.is_numeric(row, col)
        )
        self._chunks.append(masked)
        text = masked.text or None
        flat = flat_text(text)
        if text is not None and flat != text:
            self._raw_cells[sheet.locator(row, col)] = text
        return flat

    def literal(self, value: Any) -> str | None:
        """셀이 아닌 문자열(각주·라벨 등)도 같은 경로로 마스킹한다."""
        if value is None:
            return None
        masked = self.masker.mask(value)
        self._chunks.append(masked)
        return masked.text or None

    def reset(self) -> None:
        self._chunks = []
        self._raw_cells = {}

    # -- 레코드 만들기 ---------------------------------------------------
    def build(
        self,
        *,
        locator: str,
        excerpt: str,
        unit: str,
        structured: dict | None = None,
        authored_at: str | None = None,
        author_name: str | None = None,
        lang: str = "ko",
    ) -> dict:
        raw_len = len(excerpt)
        # **여기서 자르지 않는다.** 500자를 넘으면 `split_overflow` 가 `#n` 조각으로 나눠
        # 전량을 남긴다. 예전에는 `excerpt[:EXCERPT_MAX]` 로 잘라 뒤쪽을 버렸고, 그 사실이
        # `raw_len` 말고는 어디에도 남지 않았다(실측 손실: `6월23일(화)!C39` 748자 중 248자).
        text = excerpt
        rec = {
            "evidence_id": make_evidence_id(self.source_id, locator),
            "source_id": self.source_id,
            "locator": locator,
            "unit": unit,
            "excerpt": text,
            "excerpt_hash": alpha_digest(text, 32),
            "pii_masked": True,
            "pii_hits": merge_hits(self._chunks),
            "raw_len": raw_len,
            "lang": lang,
            "authored_at": authored_at,
            "author_name": author_name,
        }
        if structured is not None:
            if self._raw_cells:
                structured = {**structured, "raw_cells": dict(self._raw_cells)}
            rec["structured"] = structured
        self.reset()
        return rec


def money_cell(sheet: Sheet, eb: EvidenceBuilder, row: int, col: int | None) -> str | None:
    """금액 셀은 자릿수 구분을 넣어 읽는다.

    `1636924000` 처럼 8자리 이상 연속 숫자는 전화번호·주민번호 정규식에 그대로 걸려
    마스킹 게이트를 빨간불로 만든다(실측 4건). 값 자체는 바뀌지 않는다.
    """
    if not col:
        return None
    text = eb.cell(sheet, row, col)
    if text is None or not sheet.is_numeric(row, col):
        return text
    try:
        return f"{int(float(text)):,}"
    except ValueError:
        return text


def split_overflow(records: Iterable[dict]) -> list[dict]:
    """500자를 넘는 발췌를 `#n` 조각으로 나눈다. 원문은 한 글자도 버리지 않는다.

    계약(`contracts/parsed_record.schema.json`)이 발췌 하나를 500자로 제한하는 것은
    **한 조각의 길이**를 정하는 것이고 원본을 버리라는 뜻이 아니다. 문서 파서는 같은 상황에서
    이미 이렇게 한다(`doc_common.build_evidence`). 엑셀도 같은 관례를 쓴다.

      · 첫 조각은 원래 위치 표기와 id 를 그대로 쓴다. 기존 발췌를 가리키던 것이 안 끊긴다
      · 이어지는 조각은 `<위치>#2`, `#3` … 이고 id 를 그 위치로 다시 만든다
      · 개인정보 탐지 통계와 구조화 판정은 **첫 조각에만** 둔다. 이어지는 조각에 함께 실으면
        같은 칸을 두 번 세게 된다
      · `raw_len` 은 모든 조각에 원본 길이로 남아, 나중에 손실 여부를 셀 수 있다

    **첫 조각은 예전에 잘라 두던 것과 한 글자도 다르지 않게 둔다**(`excerpt[:500]`). 그래서
    이미 적재된 그래프에 다시 넣으면 첫 조각은 키가 같아 그대로 붙고, 버려졌던 뒤쪽만 새로
    들어온다. 재적재가 **덧붙이기**로 끝나는 것이다.

    낱말 경계를 찾아 첫 조각을 몇 글자 줄이면 읽기는 좋아지지만, 그 조각의 natural key 가
    바뀌어 이미 들어가 있는 노드와 다른 것이 되고 `evidence_id` 유일성 제약에 걸린다
    (실측으로 겪었다: `ConstraintValidationFailed` 로 적재가 멈췄다). 읽기 좋게 하려고
    이미 쌓인 것을 흔들지 않는다.
    """
    out: list[dict] = []
    for record in records:
        excerpt = str(record.get("excerpt") or "")
        if len(excerpt) <= EXCERPT_MAX:
            out.append(record)
            continue

        parts = _split_text(excerpt, EXCERPT_MAX)
        source_id = str(record.get("source_id") or "")
        base = str(record.get("locator") or "")
        for index, part in enumerate(parts, start=1):
            piece = dict(record)
            locator = base if index == 1 else f"{base}#{index}"
            piece["locator"] = locator
            piece["evidence_id"] = make_evidence_id(source_id, locator)
            piece["excerpt"] = part
            piece["excerpt_hash"] = alpha_digest(part, 32)
            piece["raw_len"] = len(excerpt)
            if index > 1:
                piece["pii_hits"] = []
                # 키는 남기고 값만 비운다. 없애면 `e["structured"]` 로 읽는 쪽이 깨진다.
                if "structured" in piece:
                    piece["structured"] = None
            out.append(piece)
    return out


def _split_text(text: str, limit: int) -> list[str]:
    """상한 글자마다 끊는다. 이어 붙이면 원문이 그대로 나온다.

    낱말 중간에서 끊길 수 있다. 그래도 이렇게 하는 이유는 `split_overflow` 의 설명에 있다:
    첫 조각을 예전과 똑같이 두어야 이미 적재된 그래프에 덧붙일 수 있다.
    """
    return [text[i : i + limit] for i in range(0, len(text), limit)]


def join_excerpt(parts: Iterable[Any], sep: str = " | ") -> str:
    return sep.join(str(p) for p in parts if p not in (None, ""))


def compose_excerpt(
    lead: Iterable[Any] = (),
    extras: Iterable[tuple[str, Any]] = (),
    sep: str = " | ",
) -> str:
    """행 발췌를 만든다. lead 는 라벨 없이, extras 는 `라벨: 값` 으로 뒤에 붙인다.

    헤더 아래 값이 있는 컬럼은 둘 중 하나에 반드시 들어가야 한다.
    structured 에만 담긴 값은 Evidence fulltext 검색에서 영영 안 잡힌다.
    """
    parts = [flat_text(v) for v in lead]
    for label, value in extras:
        text = flat_text(value)
        if not text:
            continue
        label_text = flat_text(label)
        parts.append(f"{label_text}: {text}" if label_text else text)
    return sep.join(p for p in parts if p)


def head_block_evidence(
    sheet: Sheet,
    eb: EvidenceBuilder,
    *,
    last_row: int,
    last_col: int,
    section: str,
    record_kind: str = "doc_section",
) -> list[dict]:
    """헤더 행 위에 있는 제목·안내문 블록을 Evidence 1건으로 남긴다.

    표 파서는 헤더 아래만 읽으므로 시트 제목·작성 룰·문서 메타가 통째로 사라진다.
    값이 없으면 빈 목록을 돌려준다.
    """
    eb.reset()
    parts: list[str] = []
    seen: set[str] = set()
    first: tuple[int, int] | None = None
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            raw = sheet.raw(row, col)
            if isinstance(raw, (_dt.datetime, _dt.date)):
                # 머리말의 날짜는 '2026-08-05 00:00:00' 이 아니라 사람이 읽는 ISO 로 적는다.
                iso, status = parse_cell_date(raw)
                text = eb.literal(iso) if status == "ok" else None
            else:
                text = eb.cell(sheet, row, col)
            if not text:
                continue
            if first is None:
                first = (row, col)
            # 가로·세로 병합 때문에 같은 값이 여러 칸에서 반복된다. 한 번만 적는다.
            if text in seen:
                continue
            seen.add(text)
            parts.append(text)
    if first is None:
        eb.reset()
        return []
    return [
        eb.build(
            locator=sheet.locator(*first),
            excerpt=join_excerpt(parts),
            unit="row",
            structured={
                "record_kind": record_kind,
                "sheet": sheet.title,
                "row": first[0],
                "섹션": section,
            },
        )
    ]


# --------------------------------------------------------------------------
# SourceRecord
# --------------------------------------------------------------------------
def build_source_record(
    source_id: str,
    *,
    parser: str,
    evidence_count: int,
    version: str | None = None,
    author_role: str | None = None,
    notes: str | None = None,
) -> dict:
    cfg = source_config(source_id)
    path = Path(cfg["path"])
    stat = path.stat()
    return {
        "source_id": source_id,
        "source_type": cfg["source_type"],
        "format": cfg["format"],
        "canonical_location": str(path),
        "version": version,
        "content_hash": sha256_file(path),
        "modified_at": _dt.datetime.fromtimestamp(stat.st_mtime, _dt.timezone.utc)
        .isoformat()
        .replace("+00:00", "Z"),
        "origin": "internal",
        "author_role": author_role or cfg.get("origin"),
        "doc_status": DOC_STATUS_MAP.get(cfg.get("doc_status") or "", None),
        "acl_visibility": cfg["visibility"],
        "sensitivity": SENSITIVITY_MAP[cfg["sensitivity"]],
        "pii_flag": bool(cfg["pii_flag"]),
        "source_of_record_for": [cfg["source_of_record_for"]],
        "parser": parser,
        "parser_version": PARSER_VERSION,
        "parsed_at": _dt.datetime.now(_dt.timezone.utc).isoformat().replace("+00:00", "Z"),
        "evidence_count": evidence_count,
        "extractor": "openpyxl",
        "notes": notes,
    }


# --------------------------------------------------------------------------
# 출력
# --------------------------------------------------------------------------
def write_result(result: ParseResult, out_dir: Path | None = None) -> tuple[Path, Path]:
    out = out_dir or PARSED_DIR
    out.mkdir(parents=True, exist_ok=True)
    sid = result.source["source_id"]
    src_path = out / f"{sid}.source.json"
    jsonl_path = out / f"{sid}.jsonl"
    src_path.write_text(
        json.dumps(result.source, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    with open(jsonl_path, "w", encoding="utf-8") as fh:
        for rec in result.evidence:
            fh.write(json.dumps(rec, ensure_ascii=False, sort_keys=True) + "\n")
    return src_path, jsonl_path
