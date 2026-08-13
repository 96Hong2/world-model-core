"""영업 활동일지 파서 — `영업본부_방문 및 접촉 활동 일지.xlsx` 형태의 sales_activity_log.

한 파일 안에 성격이 다른 스키마가 4벌 들어 있어 서브파서를 나눠 자동 판별한다.
  사람별 파이프라인 4시트(hidden) / 날짜별 접촉 일지 2시트 / 일일보고 14시트(hidden) / 매출전망 1시트

⚠️ 21시트 중 19시트가 hidden 이다. 숨김을 건너뛰면 딜 사실이 통째로 사라진다.
일일보고 14장은 서로 복사본이 아니라 누적 스냅샷이라, 값이 바뀐 시점만 Evidence 로 남긴다.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl.utils import get_column_letter

from ingestion.parsers.xlsx_common import (
    EvidenceBuilder,
    ParseResult,
    Sheet,
    alpha_digest,
    build_source_record,
    compose_excerpt,
    find_col,
    head_block_evidence,
    header_map,
    join_excerpt,
    load_workbook,
    money_cell,
    parse_cell_date,
    source_config,
    split_overflow,
)

SOURCE_ID = "src_sales_activity_log"
PARSER = "xlsx_sales_activity"

# 사람별 시트 이름. 도메인 프로필의 실제 담당자 시트명으로 바꿔 쓴다 (여기는 합성 예시).
PERSON_SHEETS = ("홍길동", "김철수", "이영희", "박민수")
DATE_SHEET_RE = re.compile(r"^\d+월\d+일~\d+일$")
DAILY_SHEET_RE = re.compile(r"^(\d+)월(\d+)일(\(.\))?$")
# 회사 매출 집계 시트. 딜 사실이 아니라 축이 달라 서브파서를 따로 둔다(A-10 으로 1A 에 포함).
FORECAST_SHEETS = ("매출전망",)

PERSON_HEADER_ROW = 7  # r7:r8 세로 병합이라 2줄처럼 보이지만 r8 은 전부 공백이다
PERSON_FFILL = (3, 4, 6)  # 영업대표 · 업권 · 고객사
DATE_HEADER_ROW = 4
DATE_FFILL = (2, 3, 4)  # 일자 · 담당자 · 고객사

SECTION_RE = re.compile(r"^\s*(\d+)\.\s*(.+?)\s*$")
CONTACT_PERSON_COLUMN = "고객사 담당자"


def _daily_sort_key(name: str) -> tuple[int, int]:
    m = DAILY_SHEET_RE.match(name)
    return (int(m.group(1)), int(m.group(2))) if m else (99, 99)


def _section_rows(sheet: Sheet, last_col: int = 20) -> dict[int, tuple[int, str]]:
    """일일보고 시트의 섹션 헤더 행을 찾는다. 섹션 제목과 컬럼 헤더가 같은 행에 있다."""
    out: dict[int, tuple[int, str]] = {}
    for row in range(1, min(sheet.max_row, 120) + 1):
        for c in (2, 3):
            v = sheet.text(row, c)
            if not v:
                continue
            m = SECTION_RE.match(v)
            if m and m.group(1).isdigit():
                out[int(m.group(1))] = (row, m.group(0).strip())
                break
    return out



def _parse_person_sheet(sheet: Sheet, eb: EvidenceBuilder, result: ParseResult) -> list[dict]:
    hdr = header_map(sheet, PERSON_HEADER_ROW, last_col=22)
    col = {k: find_col(hdr, k) for k in (
        "번호", "영업대표", "업권", "업무 도메인", "고객사", "단계", "최초인입일",
        "최종활동일", "활동내역", "활동유형", "상태(Status)", "현재이슈",
        "고객 니즈", "수주 가능성", "예상 매출 금액", "예상 수주월",
        "Next Action", "Action 일정", "내부지원사항", "비고",
    )}
    cols = [c for c in col.values() if c]
    out = []
    for row in range(PERSON_HEADER_ROW + 2, sheet.max_row + 1):
        if not sheet.row_has_own_value(row, cols):
            continue
        eb.reset()
        first_seen, s1 = parse_cell_date(sheet.raw(row, col["최초인입일"]))
        last_seen, s2 = parse_cell_date(sheet.raw(row, col["최종활동일"]))
        for s in (s1, s2):
            if s == "unparsable":
                result.add_failure()
        get = lambda k: eb.cell(sheet, row, col[k]) if col[k] else None  # noqa: E731
        structured = {
            "record_kind": "deal_row",
            "스키마": "사람별",
            "sheet": sheet.title,
            "row": row,
            "번호": get("번호"),
            "영업대표": get("영업대표"),
            "업권": get("업권"),
            "업무_도메인": get("업무 도메인"),
            "고객사": get("고객사"),
            "단계": get("단계"),
            "최초인입일": first_seen,
            "최종활동일": last_seen,
            "활동내역": get("활동내역"),
            "활동유형": get("활동유형"),
            "상태": get("상태(Status)"),
            "현재이슈": get("현재이슈"),
            "고객_니즈": get("고객 니즈"),
            "수주_가능성": get("수주 가능성"),
            "예상_매출_금액": money_cell(sheet, eb, row, col["예상 매출 금액"]),
            "예상_수주월": get("예상 수주월"),
            "next_action": get("Next Action"),
            "action_일정": get("Action 일정"),
            "내부지원사항": get("내부지원사항"),
            "비고": get("비고"),
        }
        # 앞쪽은 사람이 먼저 읽는 딜 문장, 뒤쪽은 라벨 붙인 나머지 컬럼이다.
        # 최초인입일·최종활동일은 "언제 기준인가"를 답할 유일한 근거인데
        # structured 에만 있으면 발췌 검색에서 사라진다.
        label = lambda k: (sheet.flat(PERSON_HEADER_ROW, col[k]) or k) if col[k] else k  # noqa: E731
        excerpt = compose_excerpt(
            [
                structured["고객사"], structured["업권"], structured["업무_도메인"],
                structured["단계"], structured["활동유형"], structured["활동내역"],
                structured["현재이슈"],
            ],
            [
                (label("최초인입일"), structured["최초인입일"]),
                (label("최종활동일"), structured["최종활동일"]),
                (label("영업대표"), structured["영업대표"]),
                (label("상태(Status)"), structured["상태"]),
                (label("고객 니즈"), structured["고객_니즈"]),
                (label("수주 가능성"), structured["수주_가능성"]),
                (label("예상 매출 금액"), structured["예상_매출_금액"]),
                (label("예상 수주월"), structured["예상_수주월"]),
                (label("Next Action"), structured["next_action"]),
                (label("Action 일정"), structured["action_일정"]),
                (label("내부지원사항"), structured["내부지원사항"]),
                (label("비고"), structured["비고"]),
                (label("번호"), structured["번호"]),
            ],
        )
        out.append(
            eb.build(
                locator=sheet.locator(row, col["번호"] or 2),
                excerpt=excerpt,
                unit="row",
                structured=structured,
                authored_at=f"{last_seen}T00:00:00Z" if last_seen else None,
            )
        )
    return out


def _parse_date_sheet(sheet: Sheet, eb: EvidenceBuilder, result: ParseResult) -> list[dict]:
    hdr = header_map(sheet, DATE_HEADER_ROW, last_col=12)
    col = {k: find_col(hdr, k) for k in (
        "일자", "담당자", "고객사", "접촉방식", "미팅내용",
        "고객사 반응 및 주요사항", "고객사 담당자", "연락처",
    )}
    cols = [c for c in col.values() if c]
    out = []
    for row in range(DATE_HEADER_ROW + 1, sheet.max_row + 1):
        if not sheet.row_has_own_value(row, cols):
            continue
        eb.reset()
        day, status = parse_cell_date(sheet.raw(row, col["일자"]))
        if status == "unparsable":
            result.add_failure()
        contact_person = eb.cell(
            sheet, row, col["고객사 담당자"], column_name=CONTACT_PERSON_COLUMN
        ) if col["고객사 담당자"] else None
        phone = eb.cell(sheet, row, col["연락처"]) if col["연락처"] else None
        structured = {
            "record_kind": "activity_row",
            "스키마": "날짜별",
            "sheet": sheet.title,
            "row": row,
            "일자": day,
            "담당자": eb.cell(sheet, row, col["담당자"]),
            "고객사": eb.cell(sheet, row, col["고객사"]),
            "접촉방식": eb.cell(sheet, row, col["접촉방식"]),
            "미팅내용": eb.cell(sheet, row, col["미팅내용"]),
            "고객사_반응": eb.cell(sheet, row, col["고객사 반응 및 주요사항"]),
            "고객사_담당자": contact_person,
            "연락처": phone,
        }
        # 담당자·고객사 담당자·연락처는 값이 마스킹 대상이라도 컬럼이 있다는 사실 자체가
        # "연락처를 아는가"를 묻는 질문의 답이다. 라벨과 마스킹된 값을 함께 남긴다.
        label = lambda k: (sheet.flat(DATE_HEADER_ROW, col[k]) or k) if col[k] else k  # noqa: E731
        excerpt = compose_excerpt(
            [
                day, structured["고객사"], structured["접촉방식"],
                structured["미팅내용"], structured["고객사_반응"],
            ],
            [
                (label("담당자"), structured["담당자"]),
                (label("고객사 담당자"), structured["고객사_담당자"]),
                (label("연락처"), structured["연락처"]),
            ],
        )
        out.append(
            eb.build(
                locator=sheet.locator(row, col["일자"] or 2),
                excerpt=excerpt,
                unit="row",
                structured=structured,
                authored_at=f"{day}T00:00:00Z" if day else None,
                author_name=structured["담당자"],
            )
        )
    return out


def _parse_daily_sheet(sheet: Sheet, eb: EvidenceBuilder, result: ParseResult) -> list[dict]:
    """일일보고 시트에서 행 구조를 가진 섹션(2·5·6)만 뽑는다.

    1·3·4 섹션은 사람별 KPI 집계표라 딜 사실이 아니다.
    Next Action 컬럼이 시트마다 M/N 으로 이동하므로 섹션 헤더 행에서 매번 다시 읽는다.
    """
    sections = _section_rows(sheet)
    out: list[dict] = []
    if not sections:
        return out
    ordered = sorted(sections.items(), key=lambda kv: kv[1][0])
    bounds = {}
    for i, (num, (row, title)) in enumerate(ordered):
        end = ordered[i + 1][1][0] if i + 1 < len(ordered) else sheet.max_row + 1
        bounds[num] = (row, title, end)

    specs = {
        2: ("deal_row", "일일보고_계약", ("고객사", "금액", "담당", "비고")),
        5: ("deal_row", "일일보고_영업진행",
            ("고객사", "단계", "예상매출", "성공확률", "현재 이슈", "Next Action", "지원 필요")),
        6: ("activity_row", "일일보고_전략BD",
            ("고객사", "담당자", "연락처", "영업대표", "상세내용", "비고")),
    }
    for num, (kind, schema, labels) in specs.items():
        if num not in bounds:
            continue
        head_row, title, end = bounds[num]
        hdr = header_map(sheet, head_row, last_col=20)
        col = {lab: find_col(hdr, lab) for lab in labels}
        if num == 5 and col.get("Next Action"):
            result.stats.setdefault("일일보고_next_action_컬럼", {})[sheet.title] = col["Next Action"]
        account_col = col.get("고객사")
        if not account_col:
            continue
        for row in range(head_row + 1, end):
            if not sheet.text(row, account_col):
                continue
            if num == 5:
                result.stats["일일보고_5번_원시행"] = result.stats.get("일일보고_5번_원시행", 0) + 1
            eb.reset()
            values = {}
            for lab, c in col.items():
                if not c:
                    values[lab] = None
                    continue
                if lab in {"금액", "예상매출"}:
                    values[lab] = money_cell(sheet, eb, row, c)
                    continue
                column_name = CONTACT_PERSON_COLUMN if lab == "담당자" and num == 6 else None
                values[lab] = eb.cell(sheet, row, c, column_name=column_name)
            structured = {
                "record_kind": kind,
                "스키마": schema,
                "sheet": sheet.title,
                "row": row,
                "섹션": title,
                "고객사": values.get("고객사"),
            }
            for lab, v in values.items():
                if lab == "고객사":
                    continue
                structured[lab.replace(" ", "_")] = v
            # 값이 바뀔 때만 남기기 위한 서명. 시트·행은 서명에 넣지 않는다.
            structured["value_signature"] = alpha_digest(
                "|".join(str(values.get(lab) or "") for lab in labels)
            )
            excerpt = join_excerpt([values.get(lab) for lab in labels])
            out.append(
                eb.build(
                    locator=sheet.locator(row, account_col),
                    excerpt=excerpt,
                    unit="row",
                    structured=structured,
                )
            )
    return out


def _parse_forecast_sheet(sheet: Sheet, eb: EvidenceBuilder) -> list[dict]:
    """매출전망 시트. 2줄 헤더가 가로 병합이라 왼쪽 값을 오른쪽으로 이어 라벨을 만든다.

    사람별·날짜별과 축이 완전히 달라(딜이 아니라 회사 매출 집계) 별도 서브파서로 둔다.
    """
    top_row, sub_row, first_data = 3, 4, 5
    last_col = min(sheet.ws.max_column or 1, 14)
    labels: dict[int, str] = {}
    carried: str | None = None
    for c in range(2, last_col + 1):
        top = sheet.text(top_row, c)
        if top:
            carried = top
        sub = sheet.text(sub_row, c)
        parts = [p for p in (carried, sub) if p]
        labels[c] = " ".join(parts) if parts else get_column_letter(c)

    def money(row: int, col: int) -> str | None:
        return money_cell(sheet, eb, row, col)

    out: list[dict] = []
    for row in range(first_data, sheet.max_row + 1):
        eb.reset()
        values = {c: money(row, c) for c in range(2, last_col + 1)}
        if not any(values.values()):
            continue
        # 왼쪽 분류 칸(B~D)은 라벨이 아니라 값이다. 앞에 그대로 두어 문맥을 만든다.
        head = join_excerpt([values.get(c) for c in (2, 3, 4)], sep=" ")
        body = join_excerpt(
            [f"{labels[c]}: {values[c]}" for c in range(5, last_col + 1) if values.get(c)]
        )
        first_col = next(c for c in range(2, last_col + 1) if values.get(c))
        out.append(
            eb.build(
                locator=sheet.locator(row, first_col),
                excerpt=join_excerpt([head, body]),
                unit="row",
                structured={
                    "record_kind": "deal_row",
                    "스키마": "매출전망",
                    "sheet": sheet.title,
                    "row": row,
                    "구분": head or None,
                    "값": {labels[c]: v for c, v in values.items() if v and c >= 5},
                },
            )
        )
    return out


def _dedup_snapshots(records: list[dict]) -> list[dict]:
    """같은 딜의 같은 값이 여러 시트에 반복되면 처음 관측만 남기고 나머지는 접는다."""
    kept: dict[tuple, dict] = {}
    order: list[tuple] = []
    for rec in records:
        st = rec["structured"]
        key = (st["스키마"], st.get("고객사"), st["value_signature"])
        if key in kept:
            kept[key]["structured"]["관측_시트"].append(st["sheet"])
            continue
        st["관측_시트"] = [st["sheet"]]
        kept[key] = rec
        order.append(key)
    return [kept[k] for k in order]


def parse(path: Path | None = None, *, daily_sheets: str = "all") -> ParseResult:
    """daily_sheets: 'all' 이면 일일보고 14장 전부를 읽고 값 변화만 남긴다.

    sources.yaml 의 scope 는 최신 2장만 넣도록 적어 두었으나, 그 2장은 5·6 섹션이
    헤더만 있고 데이터가 0행이다(실측). 그대로 따르면 딜 사실이 하나도 안 나온다.
    """
    cfg = source_config(SOURCE_ID)
    src_path = Path(path or cfg["path"])
    wb = load_workbook(src_path)

    eb = EvidenceBuilder(SOURCE_ID)
    result = ParseResult(source={}, stats={"date_parse_failures": 0, "일일보고_5번_원시행": 0})

    configured = set(cfg.get("scope", {}).get("sheets_include", []))
    daily_records: list[dict] = []

    for name in wb.sheetnames:
        if name in FORECAST_SHEETS:
            sheet = Sheet(wb[name])
            # 2줄 헤더(r3·r4) 위에 기간·전일대비 증감 같은 문맥이 있다. 데이터 행만 읽으면 사라진다.
            result.evidence += head_block_evidence(
                sheet, eb, last_row=4, last_col=14, section="시트 머리말"
            )
            result.evidence += _parse_forecast_sheet(sheet, eb)
        elif name in PERSON_SHEETS:
            sheet = Sheet(wb[name], PERSON_FFILL)
            # 헤더 위 머리말 블록에 시트 제목과 '작성 룰' 4개가 들어 있다.
            # 표 파서는 헤더 아래만 읽으므로 이 블록을 따로 남기지 않으면 통째로 사라진다.
            result.evidence += head_block_evidence(
                sheet, eb, last_row=PERSON_HEADER_ROW - 1, last_col=22, section="시트 머리말"
            )
            result.evidence += _parse_person_sheet(sheet, eb, result)
        elif DATE_SHEET_RE.match(name):
            sheet = Sheet(wb[name], DATE_FFILL)
            result.evidence += head_block_evidence(
                sheet, eb, last_row=DATE_HEADER_ROW - 1, last_col=12, section="시트 머리말"
            )
            result.evidence += _parse_date_sheet(sheet, eb, result)
        elif DAILY_SHEET_RE.match(name):
            if daily_sheets == "config" and name not in configured:
                continue
            daily_records.append((name, Sheet(wb[name])))
        else:
            result.warnings.append(f"분류하지 못한 시트: {name}")

    daily_evidence: list[dict] = []
    for name, sheet in sorted(daily_records, key=lambda x: _daily_sort_key(x[0])):
        daily_evidence += _parse_daily_sheet(sheet, eb, result)
    result.evidence += _dedup_snapshots(daily_evidence)
    result.stats["일일보고_시트수"] = len(daily_records)
    result.stats["일일보고_dedup_전"] = len(daily_evidence)
    wb.close()

    # 500자를 넘는 칸은 잘라 버리지 않고 `#n` 조각으로 나눈다. 자료 수를 세기 전에 한다.
    result.evidence = split_overflow(result.evidence)

    result.source = build_source_record(
        SOURCE_ID,
        parser=PARSER,
        evidence_count=len(result.evidence),
        notes=(
            "hidden 시트 포함. 일일보고 14장은 누적 스냅샷이라 값이 바뀐 시점만 Evidence 로 남겼다. "
            "매출전망 시트는 축이 달라 별도 서브파서로 읽는다."
        ),
    )
    return result


if __name__ == "__main__":  # pragma: no cover
    from ingestion.parsers.xlsx_common import write_result

    r = parse()
    print(write_result(r), len(r.evidence), r.stats)
